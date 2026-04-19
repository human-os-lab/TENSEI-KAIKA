from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, JSONResponse
import os
import json
from datetime import datetime
from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import RealDictCursor
import anthropic

app = FastAPI()

EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'kantei_db.xlsx')
PASSWORD = 'Lamat'

def get_db_url():
    url = os.environ.get('DATABASE_URL', '')
    # postgresql:// を postgres:// に変換（psycopg2対応）
    if url.startswith('postgres://'):
        url = url.replace('postgres://', 'postgresql://', 1)
    return url

def get_conn():
    return psycopg2.connect(get_db_url())

def init_db():
    conn = get_conn()
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS records (
            id SERIAL PRIMARY KEY,
            created_at TEXT,
            name TEXT,
            birthday TEXT,
            comment TEXT,
            result TEXT
        )
    ''')
    conn.commit()
    conn.close()

def save_record(name, birthday, comment, result):
    conn = get_conn()
    c = conn.cursor()
    c.execute(
        'INSERT INTO records (created_at, name, birthday, comment, result) VALUES (%s, %s, %s, %s, %s) RETURNING id',
        (datetime.now().strftime('%Y-%m-%d %H:%M'), name, birthday, comment, result)
    )
    record_id = c.fetchone()[0]
    conn.commit()
    conn.close()
    return record_id

def get_records():
    conn = get_conn()
    c = conn.cursor(cursor_factory=RealDictCursor)
    c.execute('SELECT created_at, name, birthday, comment, result FROM records ORDER BY created_at DESC')
    rows = c.fetchall()
    conn.close()
    return [dict(r) for r in rows]

init_db()

def load_all_data():
    wb = load_workbook(EXCEL_PATH, read_only=True)

    ws = wb['五芒星運命の固定数']
    rows = list(ws.iter_rows(values_only=True))
    meimei_db = {}
    base_year = 1924
    for i, r in enumerate(rows[1:], 1):
        year = base_year + (i - 1) // 12
        eto = r[3]
        month = r[4]
        unmei = r[6]
        ab = r[7]
        if isinstance(unmei, int) and eto and month:
            key = f'{year}{month}'
            meimei_db[key] = {'unmei': unmei, 'eto': eto, 'ab': ab}

    ws = wb['星']
    hoshi_db = {}
    for r in ws.iter_rows(values_only=True):
        if isinstance(r[0], int):
            hoshi_db[r[0]] = {'hoshi': r[1], 'tenyoshin': r[2]}

    ws = wb['ジュメリ']
    jumeri_db = {}
    rows_j = list(ws.iter_rows(values_only=True))
    for i, r in enumerate(rows_j[1:], 0):
        eto = r[0]
        hoshi = r[1]
        ab = 'α' if i % 2 == 0 else 'β'
        if eto and hoshi:
            key = f'{eto}{hoshi}{ab}'
            jumeri_db[key] = r[3]

    return meimei_db, hoshi_db, jumeri_db

meimei_db, hoshi_db, jumeri_db = load_all_data()

SYSTEM_PROMPT = """## 役割
天性開花学の鑑定AIです。
独自の天性タイプシステムから、その人の本質・強み・課題を読み解き、自己理解のきっかけを提供します。

## 鑑定の哲学
- 答えはユーザーの内側にある。引き出すだけ
- 設計図を読む。判断しない
- 短く・本質だけ・押しつけない
- ユーザーの考える力を奪わない

## 禁止事項
- 過剰な褒め／長文の説明／同じ内容の繰り返し
- 運命の決めつけ／恐怖・不安を煽る表現
- 「必ずこうなります」という断定

## 出力スタイル
- 1セクション200字以内
- 断定しない。傾向として提示する
- 重要なものに絞る。全部出さない
- 必ずJSON形式のみで出力する（前後に説明文不要）"""

def generate_kantei_text(tensei: str) -> dict | None:
    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        return None
    try:
        client = anthropic.Anthropic(api_key=api_key)
        user_prompt = f"""天性タイプ「{tensei}」の鑑定をしてください。

以下のJSON形式のみで出力してください：
{{
  "theme": "このタイプを一言で表すキーワード（15文字以内）",
  "sections": [
    {{"tag": "NATURE", "title": "天性の姿", "text": "天性の説明（200字以内）"}},
    {{"tag": "STRENGTH", "title": "強み", "text": "強みの説明（200字以内）"}},
    {{"tag": "SHADOW", "title": "課題", "text": "課題の説明（200字以内）"}}
  ]
}}"""
        message = client.messages.create(
            model="claude-opus-4-6",
            max_tokens=1024,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": user_prompt}]
        )
        text = message.content[0].text.strip()
        # JSON部分だけ抽出（```json ... ``` ブロックも対応）
        if text.startswith('```'):
            text = text.split('```')[1]
            if text.startswith('json'):
                text = text[4:]
        return json.loads(text)
    except Exception:
        return None

def calc_kantei(birthday_str):
    try:
        from datetime import datetime as dt
        bd = dt.strptime(birthday_str, '%Y-%m-%d').date()
    except:
        return None, '生年月日の形式が正しくありません'

    year_month_key = f'{bd.year}{bd.month:02d}'
    meimei = meimei_db.get(year_month_key)
    if not meimei:
        return None, '対応範囲外の生年月日です（1924年〜2020年）'

    unmei_fixed = meimei['unmei']
    eto = meimei['eto']
    ab = meimei['ab']

    gohoshi_num = unmei_fixed + bd.day
    if gohoshi_num > 60:
        gohoshi_num -= 60

    hoshi_data = hoshi_db.get(gohoshi_num, {})
    hoshi = hoshi_data.get('hoshi', '?')

    jumeri_base = f'{eto}{hoshi}{ab}'
    jumeri = jumeri_db.get(jumeri_base, '単星')

    tensei = jumeri if 'MIX' in str(jumeri) else hoshi
    return tensei, None

@app.get('/', response_class=HTMLResponse)
async def root():
    with open(os.path.join(os.path.dirname(__file__), 'index.html'), encoding='utf-8') as f:
        return f.read()

@app.post('/kantei')
async def kantei(request: Request):
    body = await request.json()
    name = body.get('name', '').strip()
    birthday = body.get('birthday', '')
    comment = body.get('comment', '')

    if not name:
        return JSONResponse({'error': 'お名前を入力してください'})
    if not birthday:
        return JSONResponse({'error': '生年月日を入力してください'})

    result, error = calc_kantei(birthday)
    if error:
        return JSONResponse({'error': error})

    kantei_text = generate_kantei_text(result)
    record_id = save_record(name, birthday, comment, result)
    return JSONResponse({'result': result, 'record_id': record_id, 'kantei_text': kantei_text})

@app.get('/history')
async def history(pw: str = ''):
    if pw != PASSWORD:
        return JSONResponse({'error': 'パスワードが違います'})
    records = get_records()
    return JSONResponse({'records': records})

@app.post('/more')
async def more(request: Request):
    body = await request.json()
    record_id = body.get('record_id')
    comment = body.get('comment', '')
    if comment and record_id:
        conn = get_conn()
        c = conn.cursor()
        c.execute(
            "UPDATE records SET comment = comment || ' 【追加質問】' || %s WHERE id = %s",
            (comment, record_id)
        )
        conn.commit()
        conn.close()
    return JSONResponse({'ok': True})
